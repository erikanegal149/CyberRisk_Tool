# data/data_manager.py
# Сохранение и загрузка данных через Excel

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# __file__ — путь к data_manager.py, поднимаемся на уровень выше в корень проекта
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(_BASE_DIR, "..", "cyberrisk_data.xlsx")


def save_data(controls, risks, settings):
    """Сохраняет все данные в рабочий Excel файл"""
    wb = Workbook()

    # ─── Лист 1: Настройки ───
    ws1 = wb.active
    ws1.title = "Настройки"
    ws1["A1"] = "org_name"
    ws1["B1"] = settings.get("org_name", "")
    ws1["A2"] = "industry"
    ws1["B2"] = settings.get("industry", "")
    ws1["A3"] = "responsible"
    ws1["B3"] = settings.get("responsible", "")
    ws1["A4"] = "date"
    ws1["B4"] = settings.get("date", "")

    # ─── Лист 2: Контроли ───
    ws2 = wb.create_sheet("Контроли")
    headers = ["id", "category", "name", "description", "standard", "score"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)

    for row_idx, control in enumerate(controls, 2):
        ws2.cell(row=row_idx, column=1, value=control["id"])
        ws2.cell(row=row_idx, column=2, value=control["category"])
        ws2.cell(row=row_idx, column=3, value=control["name"])
        ws2.cell(row=row_idx, column=4, value=control["description"])
        ws2.cell(row=row_idx, column=5, value=control["standard"])
        ws2.cell(row=row_idx, column=6, value=control["score"])

    # ─── Лист 3: Риски ───
    ws3 = wb.create_sheet("Риски")
    risk_headers = ["id", "name", "description", "probability",
                    "impact", "level", "level_text", "status"]
    for col, h in enumerate(risk_headers, 1):
        ws3.cell(row=1, column=col, value=h)

    for row_idx, risk in enumerate(risks, 2):
        ws3.cell(row=row_idx, column=1, value=risk["id"])
        ws3.cell(row=row_idx, column=2, value=risk["name"])
        ws3.cell(row=row_idx, column=3, value=risk.get("description", ""))
        ws3.cell(row=row_idx, column=4, value=risk["probability"])
        ws3.cell(row=row_idx, column=5, value=risk["impact"])
        ws3.cell(row=row_idx, column=6, value=risk["level"])
        ws3.cell(row=row_idx, column=7, value=risk["level_text"])
        ws3.cell(row=row_idx, column=8, value=risk["status"])

    wb.save(DATA_FILE)


def load_data():
    """Загружает данные из рабочего Excel файла"""
    if not os.path.exists(DATA_FILE):
        return None

    try:
        wb = load_workbook(DATA_FILE)

        # Настройки
        ws1 = wb["Настройки"]
        settings = {}
        for row in ws1.iter_rows(min_row=1, values_only=True):
            if row[0] and row[1]:
                settings[row[0]] = row[1]

        # Контроли
        ws2 = wb["Контроли"]
        controls = []
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0]:
                controls.append({
                    "id": row[0],
                    "category": row[1],
                    "name": row[2],
                    "description": row[3],
                    "standard": row[4],
                    "score": row[5] or 0
                })

        # Риски
        ws3 = wb["Риски"]
        risks = []
        for row in ws3.iter_rows(min_row=2, values_only=True):
            if row[0]:
                risks.append({
                    "id": row[0],
                    "name": row[1],
                    "description": row[2] or "",
                    "probability": row[3],
                    "impact": row[4],
                    "level": row[5],
                    "level_text": row[6],
                    "status": row[7]
                })

        return {
            "controls": controls,
            "risks": risks,
            "settings": settings
        }

    except Exception:
        # Файл повреждён или листы переименованы — не падаем, возвращаем умолчания
        print("Предупреждение: файл данных повреждён, загружены данные по умолчанию")
        from copy import deepcopy
        from models.risk_model import CONTROLS
        return {
            "controls": deepcopy(CONTROLS),
            "risks": [],
            "settings": {}
        }


def export_to_excel(controls, risks, settings):
    """Экспортирует красивый отчёт в отдельный Excel файл"""
    wb = Workbook()

    # Стили
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    green_fill = PatternFill("solid", fgColor="1a472a")
    yellow_fill = PatternFill("solid", fgColor="7d5a00")
    red_fill = PatternFill("solid", fgColor="7d1a1a")
    dark_fill = PatternFill("solid", fgColor="2d2d3f")

    header_font = Font(bold=True, color="FFFFFF", size=12)
    white_font = Font(color="FFFFFF", size=11)

    thin_border = Border(
        left=Side(style="thin", color="444466"),
        right=Side(style="thin", color="444466"),
        top=Side(style="thin", color="444466"),
        bottom=Side(style="thin", color="444466")
    )

    # ─── Лист 1: Сводка ───
    ws1 = wb.active
    ws1.title = "Сводка"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 40

    ws1["A1"] = "CyberRisk Assessment Tool"
    ws1["A1"].font = Font(bold=True, color="4a9eff", size=16)
    ws1["A2"] = "Отчёт по менеджменту киберрисков"
    ws1["A2"].font = Font(color="888888", size=12)

    ws1["A4"] = "Организация:"
    ws1["B4"] = settings.get("org_name", "Не указано")
    ws1["A5"] = "Отрасль:"
    ws1["B5"] = settings.get("industry", "Не указано")
    ws1["A6"] = "Ответственный:"
    ws1["B6"] = settings.get("responsible", "Не указано")
    ws1["A7"] = "Дата оценки:"
    ws1["B7"] = settings.get("date", "Не указано")

    for row in ws1["A4:B7"]:
        for cell in row:
            cell.font = white_font
            cell.fill = dark_fill
            cell.border = thin_border

    overall = sum(c["score"] for c in controls) / len(controls) if controls else 0
    ws1["A9"] = "Общий балл:"
    ws1["B9"] = f"{round(overall, 2)} / 5.0"
    ws1["A9"].font = header_font
    ws1["B9"].font = Font(
        bold=True,
        color="2ecc71" if overall >= 4 else "f39c12" if overall >= 2.5 else "e74c3c",
        size=12
    )

    # ─── Лист 2: Контроли ───
    ws2 = wb.create_sheet("18 Контролей")
    ws2.sheet_view.showGridLines = False

    headers = ["ID", "Категория", "Контроль", "Стандарт", "Описание", "Балл", "Уровень риска"]
    widths = [5, 22, 25, 18, 45, 8, 15]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.row_dimensions[1].height = 30

    for row_idx, control in enumerate(controls, 2):
        score = control["score"]
        if score >= 4:
            fill = green_fill
            level = "Низкий"
        elif score >= 2.5:
            fill = yellow_fill
            level = "Средний"
        else:
            fill = red_fill
            level = "Высокий"

        values = [
            control["id"], control["category"], control["name"],
            control["standard"], control["description"],
            control["score"], level
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.font = white_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

        ws2.row_dimensions[row_idx].height = 35

    # ─── Лист 3: Реестр рисков ───
    ws3 = wb.create_sheet("Реестр рисков")
    ws3.sheet_view.showGridLines = False

    risk_headers = ["ID", "Название", "Вероятность", "Ущерб", "Уровень", "Статус"]
    risk_widths = [5, 35, 15, 15, 15, 15]

    for i, (h, w) in enumerate(zip(risk_headers, risk_widths), 1):
        cell = ws3.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws3.column_dimensions[get_column_letter(i)].width = w

    for row_idx, risk in enumerate(risks, 2):
        values = [
            risk["id"], risk["name"], risk["probability"],
            risk["impact"], risk["level_text"], risk["status"]
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = dark_fill
            cell.font = white_font
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    # Сохраняем отчёт
    org = settings.get("org_name", "Отчет").replace(" ", "_")
    filename = f"CyberRisk_Report_{org}.xlsx"
    wb.save(filename)
    return filename
